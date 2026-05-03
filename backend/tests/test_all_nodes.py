"""One realistic test per node. Run via `pytest -q backend/tests/test_all_nodes.py`."""
from __future__ import annotations

import asyncio
from typing import Any

import pytest

from backend.engine.context import RunContext
from backend.engine.dag_runner import run_workflow
from backend.engine.registry import NODE_SPECS, all_specs


def _run_one(node_type: str, config: dict, incoming: dict[str, Any] | None = None, ctx: RunContext | None = None):
    """Invoke one node handler directly, isolated."""
    spec = NODE_SPECS[node_type]
    node = {"id": "n1", "type": node_type, "config": config, "label": node_type}
    ctx = ctx or RunContext()
    incoming = incoming or {}
    handler = spec.handler
    if asyncio.iscoroutinefunction(handler):
        return asyncio.run(handler(node, ctx, incoming))
    return handler(node, ctx, incoming)


# ─────────────────────────── Triggers ───────────────────────────
def test_manual_trigger():
    out = _run_one("manual_trigger", {})
    assert out["triggered"] is True


def test_api_trigger():
    out = _run_one("api_trigger", {"path": "/hook"})
    assert out["triggered"] is True
    assert out["path"] == "/hook"


def test_schedule():
    out = _run_one("schedule", {"cron": "*/5 * * * *"})
    assert out["scheduled"] is True
    assert out["cron"] == "*/5 * * * *"


def test_webhook_trigger():
    ctx = RunContext(alert_payload={"x": 1})
    out = _run_one("webhook_trigger", {}, ctx=ctx)
    assert out["webhook"] is True
    assert out["payload"] == {"x": 1}


# ─────────────────────────── Data ───────────────────────────
def test_csv_extract_all_datasets():
    for ds in ("leads.csv", "products.csv", "orders.csv", "employees.csv", "transactions.csv"):
        out = _run_one("csv_extract", {"source": ds})
        assert out["rowCount"] >= 15, f"{ds} returned only {out['rowCount']}"
        assert isinstance(out["rows"], list)
        assert isinstance(out["rows"][0], dict)


def test_csv_extract_with_limit():
    out = _run_one("csv_extract", {"source": "leads.csv", "limit": 5})
    assert out["rowCount"] == 5


def test_pdf_extract_default():
    out = _run_one("pdf_extract", {})
    assert out["pages"] >= 1
    assert out["rowCount"] >= 1
    assert "fullText" in out


def test_pdf_extract_named_doc():
    out = _run_one("pdf_extract", {"source": "contract.pdf"})
    assert "SERVICE AGREEMENT" in out["fullText"]


def test_db_query_inferred_source():
    out = _run_one("db_query", {"query": "SELECT * FROM leads"})
    assert out["rowCount"] == 20


def test_db_query_explicit_source():
    out = _run_one("db_query", {"query": "anything", "source": "products.csv"})
    assert out["rowCount"] == 20


def test_http_no_url_returns_error():
    out = _run_one("http", {})
    assert "error" in out
    assert out["rowCount"] == 0


# ─────────────────────────── Transform ───────────────────────────
def _leads_incoming():
    rows = _run_one("csv_extract", {"source": "leads.csv"})["rows"]
    return {"upstream": {"rows": rows, "rowCount": len(rows)}}


def test_filter_score_threshold():
    out = _run_one("filter", {"expression": "row.score >= 80"}, _leads_incoming())
    assert out["rowCount"] >= 1
    assert all(r["score"] >= 80 for r in out["rows"])


def test_filter_string_equality_js_style():
    out = _run_one("filter", {"expression": "row.stage === 'qualified'"}, _leads_incoming())
    assert all(r["stage"] == "qualified" for r in out["rows"])


def test_map_transform_rename_and_compute():
    incoming = {
        "u": {"rows": [{"qty": 2, "price": 10}, {"qty": 5, "price": 4}], "rowCount": 2}
    }
    out = _run_one("map_transform", {
        "mappings": [
            {"to": "revenue", "expression": "row.qty * row.price"},
            {"to": "old_qty", "from": "qty"},
        ]
    }, incoming)
    assert out["rows"][0]["revenue"] == 20
    assert out["rows"][1]["revenue"] == 20
    assert out["rows"][0]["old_qty"] == 2


def test_select_columns_projection():
    out = _run_one("select_columns", {"columns": "lead_id,email,score"}, _leads_incoming())
    assert set(out["rows"][0].keys()) == {"lead_id", "email", "score"}


def test_sort_descending():
    out = _run_one("sort", {"sortBy": "score", "order": "desc"}, _leads_incoming())
    scores = [r["score"] for r in out["rows"]]
    assert scores == sorted(scores, reverse=True)


def test_group_by_sum():
    orders = _run_one("csv_extract", {"source": "orders.csv"})["rows"]
    out = _run_one("group_by", {
        "groupBy": "region", "aggregateCol": "total", "aggregateFn": "sum", "alias": "revenue"
    }, {"u": {"rows": orders, "rowCount": len(orders)}})
    assert out["rowCount"] == 4  # 4 regions
    assert all("region" in r and "revenue" in r for r in out["rows"])
    assert sum(r["revenue"] for r in out["rows"]) == pytest.approx(sum(o["total"] for o in orders))


def test_group_by_count():
    leads = _run_one("csv_extract", {"source": "leads.csv"})["rows"]
    out = _run_one("group_by", {
        "groupBy": "stage", "aggregateCol": "lead_id", "aggregateFn": "count"
    }, {"u": {"rows": leads, "rowCount": len(leads)}})
    assert sum(r["count_lead_id"] for r in out["rows"]) == 20


def test_deduplicate_email():
    incoming = {"u": {"rows": [{"email": "a@x"}, {"email": "a@x"}, {"email": "b@x"}], "rowCount": 3}}
    out = _run_one("deduplicate", {"key": "email"}, incoming)
    assert out["rowCount"] == 2
    assert out["removed"] == 1


def test_join_inner():
    incoming = {
        "left": {"rows": [{"sku": "X1", "qty": 2}, {"sku": "X2", "qty": 5}], "rowCount": 2},
        "right": {"rows": [{"sku": "X1", "name": "A"}, {"sku": "X3", "name": "C"}], "rowCount": 2},
    }
    out = _run_one("join", {"leftKey": "sku", "rightKey": "sku", "joinType": "inner"}, incoming)
    assert out["rowCount"] == 1
    assert out["rows"][0]["name"] == "A"
    assert out["rows"][0]["qty"] == 2


def test_join_left_keeps_unmatched():
    incoming = {
        "l": {"rows": [{"sku": "X1"}, {"sku": "X9"}], "rowCount": 2},
        "r": {"rows": [{"sku": "X1", "name": "A"}], "rowCount": 1},
    }
    out = _run_one("join", {"leftKey": "sku", "rightKey": "sku", "joinType": "left"}, incoming)
    assert out["rowCount"] == 2


def test_data_merge_concat():
    incoming = {
        "a": {"rows": [{"x": 1}, {"x": 2}], "rowCount": 2},
        "b": {"rows": [{"x": 3}], "rowCount": 1},
    }
    out = _run_one("data_merge", {"strategy": "concat"}, incoming)
    assert out["rowCount"] == 3


def test_data_merge_union_dedupes():
    incoming = {
        "a": {"rows": [{"x": 1}, {"x": 2}], "rowCount": 2},
        "b": {"rows": [{"x": 2}, {"x": 3}], "rowCount": 2},
    }
    out = _run_one("data_merge", {"strategy": "union"}, incoming)
    assert out["rowCount"] == 3


def test_csv_output():
    out = _run_one("csv_output", {"filename": "test.csv"}, _leads_incoming())
    assert out["filename"] == "test.csv"
    assert "lead_id" in out["csv"].split("\n")[0]
    assert out["byteSize"] > 100


# ─────────────────────────── Logic ───────────────────────────
def test_condition_splits_rows():
    orders = _run_one("csv_extract", {"source": "orders.csv"})["rows"]
    out = _run_one("condition", {"expression": "row.total > 500"},
                   {"u": {"rows": orders, "rowCount": len(orders)}})
    assert out["_type"] == "condition"
    assert out["trueCount"] + out["falseCount"] == 20
    assert all(r["total"] > 500 for r in out["rows_true"])
    assert all(r["total"] <= 500 for r in out["rows_false"])


def test_router_buckets_by_label():
    orders = _run_one("csv_extract", {"source": "orders.csv"})["rows"]
    out = _run_one("router", {"expression": "row.region"},
                   {"u": {"rows": orders, "rowCount": len(orders)}})
    assert set(out["buckets"].keys()) == {"North", "East", "West", "South"}
    assert sum(len(v) for v in out["buckets"].values()) == 20


def test_loop_passthrough_with_cap():
    incoming = {"u": {"rows": [{"i": i} for i in range(50)], "rowCount": 50}}
    out = _run_one("loop", {"maxIterations": 10}, incoming)
    assert out["rowCount"] == 10


def test_pause_short():
    import time
    t0 = time.time()
    out = _run_one("pause", {"durationMs": 100}, {"u": {"rows": [{"a": 1}], "rowCount": 1}})
    assert (time.time() - t0) >= 0.09
    assert out["rowCount"] == 1


def test_code_filters_rows():
    incoming = {"u": {"rows": [{"v": 1}, {"v": 5}, {"v": 10}], "rowCount": 3}}
    out = _run_one("code", {"code": "result = [r for r in rows if r['v'] >= 5]"}, incoming)
    assert out["rowCount"] == 2


def test_code_falls_back_to_rows_var():
    incoming = {"u": {"rows": [{"v": 1}, {"v": 2}], "rowCount": 2}}
    out = _run_one("code", {"code": "rows = rows + [{'v': 99}]"}, incoming)
    assert out["rowCount"] == 3


def test_code_error_returns_input():
    incoming = {"u": {"rows": [{"v": 1}], "rowCount": 1}}
    out = _run_one("code", {"code": "this is not python"}, incoming)
    assert "error" in out


def test_function_with_input_and_prev():
    ctx = RunContext(alert_payload={"name": "Bob"})
    out = _run_one("function", {
        "code": "result = {'greeting': 'Hello ' + input['name'], 'prev_count': prevOutput.get('rowCount', 0)}"
    }, {"u": {"rows": [], "rowCount": 7}}, ctx=ctx)
    assert out["result"]["greeting"] == "Hello Bob"
    assert out["result"]["prev_count"] == 7


# ─────────────────────────── AI ───────────────────────────
def test_agent_stub_without_api_key(monkeypatch):
    monkeypatch.delenv("GOOGLE_API_KEY", raising=False)
    out = _run_one("agent", {"prompt": "Summarize", "task": "Be brief"},
                   _leads_incoming())
    assert out["stub"] is True
    assert "stub" in out["response"].lower()


def test_evaluator_pass_rate():
    incoming = {
        "u": {"rows": [{"score": 90}, {"score": 50}, {"score": 100}, {"score": 30}], "rowCount": 4}
    }
    out = _run_one("evaluator", {"criteria": "row.score >= 80", "label": "good"}, incoming)
    assert out["passed"] == 2
    assert out["failed"] == 2
    assert out["passRate"] == "50.0%"


# ─────────────────────────── Output ───────────────────────────
def test_response_with_template():
    out = _run_one("response", {"content": "All done!"})
    assert out["response"] == "All done!"


def test_response_falls_back_to_rows():
    out = _run_one("response", {}, _leads_incoming())
    assert isinstance(out["response"], list)
    assert out["rowCount"] == 20


def test_note_returns_content():
    out = _run_one("note", {"content": "Remember this"})
    assert out["note"] == "Remember this"


def test_excel_output_multi_tab():
    leads = _run_one("csv_extract", {"source": "leads.csv"})["rows"]
    products = _run_one("csv_extract", {"source": "products.csv"})["rows"]
    incoming = {
        "a": {"rows": leads, "rowCount": len(leads)},
        "b": {"rows": products, "rowCount": len(products)},
    }
    out = _run_one("excel_output", {"filename": "report.xlsx", "tabNames": "Leads,Products"}, incoming)
    assert out["tabs"] == 2
    assert out["tabNames"] == ["Leads", "Products"]
    assert out["rowsWritten"] == 40
    assert out["byteSize"] > 1000
    assert isinstance(out["base64"], str)
    # Verify it parses back as a real .xlsx
    import base64
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(base64.b64decode(out["base64"])))
    assert wb.sheetnames == ["Leads", "Products"]


# ─────────────────────────── Integrations (graceful stubs) ───────────────────────────
def test_gmail_stubs_without_secret(monkeypatch):
    monkeypatch.delenv("GMAIL_CLIENT_SECRET", raising=False)
    out = _run_one("gmail", {"to": "x@y.z", "subject": "Hi", "body": "Hello"})
    assert out["simulated"] is True
    assert out["needsIntegration"] == "gmail"


def test_slack_stubs_without_webhook(monkeypatch):
    monkeypatch.delenv("SLACK_WEBHOOK_URL", raising=False)
    out = _run_one("slack", {"channel": "#general", "message": "Hello"})
    assert out["simulated"] is True


def test_notion_stubs_without_api_key(monkeypatch):
    monkeypatch.delenv("NOTION_API_KEY", raising=False)
    out = _run_one("notion", {"databaseId": "abc", "action": "query"})
    assert out["simulated"] is True


def test_mcp_stubs_without_server(monkeypatch):
    monkeypatch.delenv("MCP_SERVER_URL", raising=False)
    out = _run_one("mcp", {"tool": "search", "params": {"q": "x"}})
    assert out["simulated"] is True


def test_github_list_repos_returns_lean_shape(monkeypatch):
    """GitHub node hits the real API but with no token returns public-only / 401 — guard for both."""
    import os
    if not os.getenv("GITHUB_TOKEN"):
        pytest.skip("No GITHUB_TOKEN — skipping live integration probe")
    out = _run_one("github", {"action": "list-repos"})
    assert "rows" in out and out["connected"] is True
    if out["rows"]:
        assert "name" in out["rows"][0]


# ─────────────────────────── Coverage guarantee ───────────────────────────
def test_every_registered_node_has_a_test():
    """Maintenance gate: ensure new nodes don't slip in untested."""
    import inspect
    import sys
    this_module = sys.modules[__name__]
    src = inspect.getsource(this_module)
    untested = []
    for type_id in NODE_SPECS:
        # accept any test_<word>_ that mentions the type_id, OR the explicit per-node test name
        if f'"{type_id}"' not in src and f"'{type_id}'" not in src:
            untested.append(type_id)
    assert not untested, f"Nodes lacking tests: {untested}"
